import os
import openpyxl 
import pandas as pd

xlfs = [x for x in os.listdir('.') if os.path.isfile(x) and os.path.splitext(x)[1] == '.xlsx']
# 将当前目录内所有xlsx文件名存放在xlfs列表中

def sheets_names():#获取所有sheet的名称存放在sh_names中
    num= len(xlfs)#返回文件列表的个数
    sh_names=[]#定义一个存放sheet名称的文件夹
    for n in range(num):
        wb = openpyxl.load_workbook(xlfs[n],read_only=True)
        sn=wb.get_sheet_names()#获取一个工作薄中工作表的名称，组成一个列表
        sn_num=len(sn)
        for i in range(sn_num):
            if sn[i] in sh_names:#判断工作表名称的列表里面是否包含了新的名称，如果包含，则什么也不做，
                pass
            else:
                sh_names.append(sn[i])#如果没有包含，则将新的名称添加到工作表名称列表中
    return sh_names

def create_new_workbook():#创建一个新的工作薄，用于存储合并数据
    sh_names=sheets_names()
    s_len=len(sh_names)
    wb=openpyxl.Workbook()#新建一个工作薄，用于存储合并数据
    ws=wb.active
    ws.title=sh_names[0]
    for m in range(1,s_len):
        wb.create_sheet(sh_names[m])#根据需要合并的工作表名称，在合并表中建立相应的sheet
        wb.save('合并表.xlsx')
       
def writer_excel():#读取每张工作表的数据，并写入合并表中
    sh_names=sheets_names()
    s_len=len(sh_names)
    num= len(xlfs)#返回文件列表的个数
    path=os.path.abspath('.')+'\合并表.xlsx'
    writer=pd.ExcelWriter(path)#pandas 写入文件的路径
    for j in range(s_len):#根据sheet的名称遍历所有工作薄里面，有sheet名相同的表格数据进行合并
        data1=pd.DataFrame()
        for n in range(num):
            wb1 = openpyxl.load_workbook(filename = xlfs[n],read_only=True)#遍历所有表格数据
            sn1=wb1.get_sheet_names()
            if sh_names[j] in sn1:#如果该工作薄里面有某个表，则将其进行合并
                data=pd.read_excel(xlfs[n],sheet_name=sh_names[j],header=None,skiprows=[0,1,2])#,skiprows跳过多少行)#将数据读取在data容器中
                if data.shape!=(0,0):
                    data['数据来源']=xlfs[n]#在读取的数据后，加1列，显示数据容器
                data1=pd.concat([data1,data])#将读取的各表数据进行拼接
        data1.to_excel(excel_writer=writer,sheet_name=sh_names[j],index=True)#,header=5)#将同一工作表的数据写入合并工作表内 
    writer.save()

if __name__ == "__main__":
    create_new_workbook()
    writer_excel()
